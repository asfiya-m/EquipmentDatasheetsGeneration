from openpyxl import load_workbook, Workbook
from io import BytesIO
import zipfile

def split_workbook_by_sheet(master_bytes_io):
    """
    Splits a populated master workbook into individual Excel files per sheet.

    Args:
        master_bytes_io (BytesIO): BytesIO stream of the final master workbook.

    Returns:
        BytesIO: ZIP archive stream containing one Excel file per sheet.
    """
    zip_buffer = BytesIO()
    wb_full = load_workbook(master_bytes_io, data_only=True)

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        for sheet_name in wb_full.sheetnames:
            ws_source = wb_full[sheet_name]

            # Create new workbook and copy sheet content
            wb_single = Workbook()
            ws_new = wb_single.active
            ws_new.title = sheet_name

            for row in ws_source.iter_rows(values_only=True):
                ws_new.append(row)

            # Save to BytesIO
            single_stream = BytesIO()
            wb_single.save(single_stream)
            single_stream.seek(0)

            # Sanitize filename
            safe_name = sheet_name.replace("/", "_").replace("\\", "_")
            zipf.writestr(f"{safe_name}.xlsx", single_stream.read())

    zip_buffer.seek(0)
    return zip_buffer