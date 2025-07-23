"""
populate_parameters.py

Populates parameter values into the master equipment datasheet
based on stream data from the SysCAD streamtable.

Handles:
✅ Explicit & implied equipment
✅ Parameters from Stream Table V
✅ Parameters from Equipment & Stream List (per equipment)
✅ Fallback logic if primary value missing
✅ stream_tag_override (only for specific equipment if defined in YAML `overrides`)

Author: Asfiya Khanam
Updated: July 2025
"""

from openpyxl import load_workbook
import pandas as pd
from io import BytesIO
from datetime import datetime
import yaml


def apply_conversion(value, convert_key):
    if convert_key == "multiply_1000":
        return value * 1000
    elif convert_key == "multiply_100":
        return value * 100
    elif convert_key == "divide_1000":
        return value / 1000
    elif convert_key is None:
        return value
    else:
        raise ValueError(f"Unknown convert key: {convert_key}")


def populate_parameters(master_file, streamtable_file, verbose=False):
    wb_master = load_workbook(master_file)

    with open("param_mapping.yaml", "r") as f:
        param_mapping = yaml.safe_load(f)

    df_streamlist = pd.read_excel(
        streamtable_file, sheet_name="Equipment & Stream List", header=None, engine="openpyxl"
    )
    df_streamtable = pd.read_excel(
        streamtable_file, sheet_name="Stream Table V", header=6, engine="openpyxl"
    )

    skipped = []

    # Stream Table V → dict
    streamtable_lookup = {}
    for idx, row in df_streamtable.iterrows():
        tag = str(row.iloc[0]).strip().lower()
        if tag:
            streamtable_lookup[tag] = row

    # Equipment & Stream List → map
    equip_stream_map = {}
    equipment_rows = df_streamlist.iloc[3:]
    for _, row in equipment_rows.iterrows():
        equip_name = str(row[0]).strip()
        outputs = [str(tag).strip() for tag in row[1:6] if pd.notna(tag)]
        inputs = [str(tag).strip() for tag in row[6:13] if pd.notna(tag)]
        equip_stream_map[equip_name] = {"outputs": outputs, "inputs": inputs}

    for sheet_name in wb_master.sheetnames:
        ws = wb_master[sheet_name]

        if sheet_name not in param_mapping:
            skipped.append(f"[SKIP] Sheet '{sheet_name}': no param_mapping defined")
            continue

        mapping = param_mapping[sheet_name]
        mapping_lc = {k.strip().lower(): v for k, v in mapping.items()}

        for col in range(4, ws.max_column + 1):
            cell = ws.cell(row=3, column=col)
            if not cell.value:
                continue
            equip_name = str(cell.value).strip()
            equip_col = cell.column

            streams_key = equip_name
            if sheet_name == "Agitator":
                if "-" not in equip_name:
                    skipped.append(f"[SKIP] {equip_name}: invalid format for implied equipment")
                    continue
                suffix = equip_name.split("-", 1)[1]
                streams_key = f"TK-{suffix}"

            if streams_key not in equip_stream_map:
                skipped.append(f"[SKIP] {equip_name}: no streams found for base '{streams_key}'")
                continue

            if verbose:
                print(f"\n✅ Equipment: {equip_name} → Sheet: {sheet_name} → Column: {equip_col}")

            for row_cells in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                param_name = str(row_cells[1].value).strip() if row_cells[1].value else ""
                param_lc = param_name.lower()

                if param_lc not in mapping_lc:
                    ws.cell(row=row_cells[0].row, column=equip_col).value = None
                    continue

                rule = mapping_lc.get(param_lc)
                if not rule:
                    ws.cell(row=row_cells[0].row, column=equip_col).value = None
                    continue

                # 🔷 Resolve rule_to_use (default + optional override for this equipment)
                rule_to_use = rule
                if isinstance(rule, dict) and "default" in rule and "overrides" in rule:
                    if equip_name in rule["overrides"]:
                        rule_to_use = {**rule["default"], **rule["overrides"][equip_name]}
                        if verbose:
                            print(f" → Using override rule for {equip_name} → {param_name}")
                    else:
                        rule_to_use = rule["default"]

                # 📄 Case 1: fixed text
                if "text" in rule_to_use:
                    ws.cell(row=row_cells[0].row, column=equip_col).value = rule_to_use["text"]
                    if verbose:
                        print(f" → Writing text '{rule_to_use['text']}' to {param_name}")
                    continue

                # 📄 Case 2: Equipment & Stream List per_equipment
                if rule_to_use.get("sheet") == "Equipment & Stream List" and rule_to_use.get("per_equipment"):
                    equip_row_match = None
                    for _, row in equipment_rows.iterrows():
                        if str(row[0]).strip().lower() == streams_key.lower():
                            equip_row_match = row
                            break

                    val = None
                    if equip_row_match is not None:
                        val = equip_row_match[rule_to_use["col_idx"] - 1]

                    if pd.notna(val):
                        ws.cell(row=row_cells[0].row, column=equip_col).value = round(float(val), 2)
                        if verbose:
                            print(f" → Writing {val} to {param_name} (from Equipment & Stream List)")
                        continue

                    # 📄 fallback logic
                    if not rule_to_use.get("fallback"):
                        skipped.append(f"[SKIP] {equip_name}: {param_name} missing in Equipment & Stream List, no fallback")
                        ws.cell(row=row_cells[0].row, column=equip_col).value = None
                        continue

                    if verbose:
                        print(f" → {param_name}: fallback to Stream Table V")
                    rule_to_use = rule_to_use["fallback_rule"]

                # 📄 Case 3: use_stream_name
                if rule_to_use.get("use_stream_name"):
                    stream_tags = equip_stream_map[streams_key][rule_to_use.get("stream_type", "output") + "s"]
                    idx = rule_to_use.get("stream_index", 0)
                    if idx >= len(stream_tags):
                        msg = f"[SKIP] {equip_name}: no stream at index {idx} for {param_name} (stream name)"
                        skipped.append(msg)
                        if verbose: print(msg)
                        ws.cell(row=row_cells[0].row, column=equip_col).value = None
                        continue
                    stream_name = stream_tags[idx]
                    ws.cell(row=row_cells[0].row, column=equip_col).value = stream_name
                    if verbose:
                        print(f" → Writing stream name '{stream_name}' to {param_name}")
                    continue

                # 📄 Case 4: Stream Table V lookup
                if "stream_tag_override" in rule_to_use:
                    stream_tag = rule_to_use["stream_tag_override"].strip().lower()
                    if verbose:
                        print(f" → Using overridden stream tag '{stream_tag}' for {param_name}")
                else:
                    stream_key = rule_to_use.get("stream_type", "output") + "s"
                    stream_tags = equip_stream_map.get(streams_key, {}).get(stream_key, [])
                    idx = rule_to_use.get("stream_index", 0)

                    if not stream_tags:
                        msg = f"[SKIP] {equip_name}: no streams found for {stream_key} for {param_name}"
                        skipped.append(msg)
                        if verbose: print(msg)
                        ws.cell(row=row_cells[0].row, column=equip_col).value = None
                        continue

                    if idx >= len(stream_tags):
                        msg = f"[SKIP] {equip_name}: no stream at index {idx} for {param_name}"
                        skipped.append(msg)
                        if verbose: print(msg)
                        ws.cell(row=row_cells[0].row, column=equip_col).value = None
                        continue

                    stream_tag = stream_tags[idx].strip().lower()
                    if verbose:
                        print(f" → Using normal stream tag '{stream_tag}' for {param_name}")

                if stream_tag not in streamtable_lookup:
                    msg = f"[SKIP] {equip_name}: stream '{stream_tag}' not found for {param_name}"
                    skipped.append(msg)
                    if verbose: print(msg)
                    ws.cell(row=row_cells[0].row, column=equip_col).value = None
                    continue

                stream_row = streamtable_lookup[stream_tag]
                val = stream_row.iloc[rule_to_use["col_idx"] - 1]
                if pd.isna(val):
                    msg = f"[SKIP] {equip_name}: NaN for {param_name} in stream '{stream_tag}' at col {rule_to_use['col_idx']}"
                    skipped.append(msg)
                    if verbose: print(msg)
                    ws.cell(row=row_cells[0].row, column=equip_col).value = None
                    continue

                result = float(val)
                result = apply_conversion(result, rule_to_use.get("convert"))
                ws.cell(row=row_cells[0].row, column=equip_col).value = round(result, 2)
                if verbose:
                    print(f" → Writing {result} to {param_name} (from Stream Table V)")

    output = BytesIO()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Master_DataSheet_ParametersPopulated_{timestamp}.xlsx"
    wb_master.save(output)
    output.seek(0)

    return output, filename, skipped
