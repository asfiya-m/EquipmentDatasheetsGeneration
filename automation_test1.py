# -*- coding: utf-8 -*-
"""
generate_master_datasheet.py

This script automates the creation of the Master Equipment datasheet file from 
'Datasheets.xlsm', which contains multiple equipment-specific sheets.

Key Features:
- Creates the Master eexcel sheet with a time stamp in the name of the sheet
- Extracts parameter name (Column C), units (Column E), and category (Column I) from each sheet.
- Groups parameters into five standardized categories:
  - SysCAD Inputs, Engineering Inputs, Lab/Pilot Inputs, Project Constant, Vendor Inputs
- Generates one sheet per equipment with:
  - Header rows (equipment name, unit count placeholder, bold column titles)
  - Parameters grouped under each category
  - Category labels merged vertically
  - Auto-fit column widths
  

Sheets without valid category data are skipped. Output is saved as 'Master_DataSheet_Generated_vi.xlsx'.

Requirements: pandas, openpyxl

Created on Thu May 29 17:15:36 2025

@author: AsfiyaKhanam
"""
from io import BytesIO
from datetime import datetime
from collections import defaultdict
import warnings
# import re

def generate_master_datasheet(uploaded_file):
    """
    Generates a Master Equipment Datasheet from the uploaded Excel file.
    Handles per-sheet column mappings for parameter name, units, and category.
    """
    warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')

    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side
    from openpyxl.utils import get_column_letter

    # --- TIMESTAMPED OUTPUT FILE ---
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    output_filename = f"Master_DataSheet_{timestamp}.xlsx"

    # --- CATEGORY MAPPING ---
    category_mapping = {
        "SysCAD": "SysCAD Inputs",
        "Engineering Input": "Engineering Inputs",
        "Lab/Pilot Value": "Lab/Pilot Inputs",
        "Project Constant": "Project Constant",
        "Vendor Input": "Vendor Inputs"
    }

    # ordered_categories = [
    #     "SysCAD Inputs",
    #     "Engineering Inputs",
    #     "Lab/Pilot Inputs",
    #     "Project Constant",
    #     "Vendor Inputs"
    # ]

    ordered_categories = [
        "Project Constant",
        "SysCAD Inputs",
        "Lab/Pilot Inputs",
        "Engineering Inputs",
        "Vendor Inputs"
    ]

    # --- PER-SHEET COLUMN MAPPING ---
    # Column indices are 0-based
    sheet_column_map = {
        "default": {"param": 2, "unit": 4, "category": 8},       # C, E, I
        "Heat Exchanger-1": {"param": 2, "unit": 11, "category": 25}  # C, L, Z
    }

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    xls = pd.ExcelFile(uploaded_file)
    sheet_names = xls.sheet_names

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for sheet_name in sheet_names:
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)

        # Get column mapping
        cols = sheet_column_map.get(sheet_name, sheet_column_map["default"])
        param_col = cols["param"]
        unit_col = cols["unit"]
        category_col = cols["category"]

        # Skip if category column invalid or empty
        if df.shape[1] <= category_col or df.iloc[:, category_col].dropna().empty:
            continue

        records = []
        for _, row in df.iterrows():
            param = row[param_col] if pd.notna(row[param_col]) else None
            unit = row[unit_col] if pd.notna(row[unit_col]) else ""
            category_raw = row[category_col] if pd.notna(row[category_col]) else None

            if param and category_raw:
                category = category_mapping.get(str(category_raw).strip(), None)
                if category:
                    records.append((category, str(param).strip(), str(unit).strip()))

        if not records:
            continue

        # Group by category
        grouped = {cat: [] for cat in ordered_categories}
        for category, param, unit in records:
            grouped[category].append((param, unit))

        # Create output sheet
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit

        ws.append([sheet_name])
        ws.append(["Number of units ="])
        header_row = ["Parameter Category", "Input Parameters", "Units"]
        ws.append(header_row)           # for row 3:headers
        ws.insert_rows(4)           # line to add simplified standard names

        ws.append(header_row)           # for row 4

        for cell in ws[3] + ws[4]:
            if cell.value:
                cell.font = Font(bold=True)
        
        current_row = 5
        column_widths = defaultdict(int)
        for col_idx, value in enumerate(header_row, start=1):
            column_widths[col_idx] = max(column_widths[col_idx], len(str(value)))

        # Write grouped data
        for category in ordered_categories:
            param_list = grouped[category]
            if not param_list:
                continue

            start_row = current_row

            first_param, first_unit = param_list[0]
            ws.cell(row=current_row, column=1, value=category)
            ws.cell(row=current_row, column=2, value=first_param)
            ws.cell(row=current_row, column=3, value=first_unit)

            column_widths[1] = max(column_widths[1], len(str(category)))
            column_widths[2] = max(column_widths[2], len(str(first_param)))
            column_widths[3] = max(column_widths[3], len(str(first_unit)))

            current_row += 1

            for param, unit in param_list[1:]:
                ws.cell(row=current_row, column=2, value=param)
                ws.cell(row=current_row, column=3, value=unit)
                column_widths[2] = max(column_widths[2], len(str(param)))
                column_widths[3] = max(column_widths[3], len(str(unit)))
                current_row += 1

            end_row = current_row - 1
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

        # Auto-fit columns
        for col_idx, width in column_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width + 4

        # Apply borders
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None or cell.coordinate in ws.merged_cells:
                    cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    print(f"✅ Master file ready for download: {output_filename}")
    return output, output_filename
