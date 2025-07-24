"""
populate_equipment_names.py
This script automates the population of equipment names in the Master Equipment datasheet file.
Populates equipment names into the Master Equipment datasheet, with support for:
- Explicit equipment mapping
- Implied Agitated Tanks from selected BP/PF/P units
- Implied Agitators for TK and Agitated Tanks

Created on Wed July 16
@author: AsfiyaKhanam
"""
from io import BytesIO
from datetime import datetime
from collections import defaultdict
import re
from openpyxl import load_workbook

# import pandas as pd

def populate_equipment_names(master_file, streamtable_file, verbose=True):
    """
    Populate equipment names from streamtable into master datasheet.
    Writes equipment names into each matching sheet starting from D3 (row 3, columns D..).
    
    Args:
        master_file (BytesIO or path): Master datasheet file (.xlsx).
        streamtable_file (BytesIO or path): Detailed streamtable file (.xlsx).
        
    Returns:
        output (BytesIO): Updated Excel file.
        filename (str): Timestamped output filename.
        skipped (list): Equipment names that could not be matched.
    """
    # Load master workbook
    wb_master = load_workbook(master_file)
    ws_streamlist = load_workbook(streamtable_file, data_only=True)["Equipment & Stream List"]

    skipped = []

    # explicit mapping
    equipment_sheet_map = {
        "TK": "DLE Tank",
        "BP_TK": "Bolted Panel Tank",
        "PF_TK": "PreFab Tank",
        "P_TK": "Poly Tank",
        "FP_PK": "Filter Press",
        "IX_PK": "Ion Exchange",
        "RO_PK": "Reverse Osmosis System",
        "S": "Clarifier",
        "E": "Heat Exchanger-1",
        "SL": "Silos",
        "F": "Media Filter"
    }

    # Agitated Tanks: only for specific units
    agitated_tank_units = {
        "P_TK-0102_Oxidation_Filter_Feed_Tank",
        "P_TK-0201_StripSolutionFeedTank",
        "P_TK-0204_DLEProductTank",
        "P_TK-0302_Barite_Filter_Feed_Tank",
        "P_TK-0305_Lime_Filter_Feed_Tank",
        "P_TK-0308_Carbonate_Filter_Feed_Tank",
        "PF_TK-0202_DLEFeedTank",
        "PF_TK-0203_DLE_Depleted_Brine_Tank",
        "PF_TK-0304_LimePptReactor",
        "PF_TK-0307_CarbonatePurificationReactor",
        "PF_TK-0301_BaritePptReactor",
        "BP_TK-0806_Wastewater_Treatment_Tank",
    }

    sheet_unit_counts = defaultdict(int)
    equipment_names = []

    for row in ws_streamlist.iter_rows(min_row=4, min_col=1, max_col=1):
        cell_value = str(row[0].value).strip()
        if cell_value:
            equipment_names.append(cell_value)

    for equip_name in equipment_names:
        # 🚫 Skip if no numeric part
        match = re.search(r"-.*?(\d+)", equip_name)
        if not match:
            if verbose:
                print(f"⚠️ Skipping {equip_name}: no numeric part in name")
            skipped.append(f"{equip_name}: no numeric part in name")
            continue

        matched = False
        equip_prefix = equip_name.split('-', 1)[0]

        for code, sheet_name in equipment_sheet_map.items():
            if equip_prefix == code:
                if sheet_name in wb_master.sheetnames:
                    ws = wb_master[sheet_name]
                    col_idx = 4
                    while ws.cell(row=3, column=col_idx).value:
                        col_idx += 1

                    # ws.insert_rows(4)

                    # Full SysCAD name
                    ws.cell(row=3, column=col_idx).value = equip_name

                    # Simplified PFD name in Row 4
                    parts = str(equip_name).split("_")
                    short_name = next((p for p in parts if "-" in p and any(c.isdigit() for c in p)), equip_name)
                    ws.cell(row=4, column=col_idx).value = short_name

                    sheet_unit_counts[sheet_name] += 1
                    matched = True
                    if verbose:
                        print(f"✅ Wrote {equip_name} → Sheet: {sheet_name} → Column: {col_idx}")
                else:
                    skipped.append(f"{equip_name}: sheet '{sheet_name}' not found")
                    matched = True
                break

        # ➕ Special case: add to Agitated Tanks if in list
        if equip_name in agitated_tank_units:
            sheet_name = "Agitated Tanks"
            if sheet_name in wb_master.sheetnames:
                ws = wb_master[sheet_name]
                col_idx = 4
                while ws.cell(row=3, column=col_idx).value:
                    col_idx += 1

                #  SysCAD name in the sheet
                ws.cell(row=3, column=col_idx).value = equip_name

                #  Short name in row 4
                parts = str(equip_name).split("_")
                short_name = next((p for p in parts if "-" in p and any(c.isdigit() for c in p)), equip_name)
                ws.cell(row=4, column=col_idx).value = short_name

                sheet_unit_counts[sheet_name] += 1
                if verbose:
                    print(f"💧 Agitated: Wrote {equip_name} → Agitated Tanks → Column: {col_idx}")
            else:
                skipped.append(f"{equip_name}: Agitated Tanks sheet not found")

            # ➕ Implied Agitator
            if "-" in equip_name:
                suffix = equip_name.split("-", 1)[1]
                agitator_tag = f"A-{suffix}"
                agitator_sheet = "Agitator"
                if agitator_sheet in wb_master.sheetnames:
                    ws_ag = wb_master[agitator_sheet]
                    col_idx = 4
                    while ws_ag.cell(row=3, column=col_idx).value:
                        col_idx += 1

                    #  SysCAD names
                    ws_ag.cell(row=3, column=col_idx).value = agitator_tag
                    #  short names
                    short_ag_name = agitator_tag.split("_")[0]
                    ws_ag.cell(row=4, column=col_idx).value = short_ag_name

                    sheet_unit_counts[agitator_sheet] += 1
                    if verbose:
                        print(f"✨ Implied: Wrote {agitator_tag} → Agitator → Column: {col_idx}")
                else:
                    skipped.append(f"{agitator_tag}: Agitator sheet not found")

        if not matched:
            skipped.append(f"{equip_name}: no mapping for code")
            if verbose:
                print(f"⚠️ {equip_name}: no mapping for code")

    # Write unit counts in B2
    for sheet_name, count in sheet_unit_counts.items():
        ws = wb_master[sheet_name]
        ws.cell(row=2, column=2).value = count
        if verbose:
            print(f"🔷 Sheet '{sheet_name}': unit count = {count} (written to B2)")

    output = BytesIO()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Master_DataSheet_EquipmentPopulated_{timestamp}.xlsx"
    wb_master.save(output)
    output.seek(0)

    return output, filename, skipped

