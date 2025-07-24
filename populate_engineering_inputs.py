"""
populate_engineering_inputs.py

Step 4: Populates *Engineering Inputs* into the master equipment datasheet
from the corresponding datasheets workbook (col K).

Features:
✅ Handles merged category cells correctly (remembers current category).
✅ Reads parameters under 'Engineering Inputs' in master sheet.
✅ Looks up each parameter in datasheets sheet (col C).
✅ Writes value from col K into all units in master.
✅ Logs skipped parameters if not found.

Author: Asfiya Khanam
Updated: July 2025
"""

from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime


def normalize(s):
    """Normalize parameter names: lower, stripped, remove trailing colon."""
    return str(s).strip().lower().rstrip(':')


def populate_engineering_inputs(master_file, datasheets_file, verbose=False):
    wb_master = load_workbook(master_file)
    wb_datasheets = load_workbook(datasheets_file, data_only=True)

    skipped = []

    for sheet_name in wb_master.sheetnames:
        ws_master = wb_master[sheet_name]

        if sheet_name not in wb_datasheets.sheetnames:
            skipped.append(f"[SKIP] Sheet '{sheet_name}' not found in datasheets workbook.")
            continue

        ws_data = wb_datasheets[sheet_name]

        if verbose:
            print(f"\n✅ Processing sheet: {sheet_name}")

        # Build parameter → value map from datasheets sheet (col C & col K)
        param_value_map = {}
        for excel_row in ws_data.iter_rows(min_row=1, max_row=ws_data.max_row):
            param_cell = excel_row[2] if len(excel_row) > 2 else None  # col C
            value_cell = excel_row[10] if len(excel_row) > 10 else None  # col K

            if param_cell is None or value_cell is None:
                continue  # skip short rows

            param = str(param_cell.value).strip() if param_cell.value else ""
            val = value_cell.value

            if not param:
                continue  # skip if param name is empty

            param_value_map[normalize(param)] = val

        max_col = ws_master.max_column
        current_category = None

        for row_cells in ws_master.iter_rows(min_row=5, max_row=ws_master.max_row, min_col=1, max_col=2):
            # Track current category (handles merged cells properly)
            if row_cells[0].value:
                current_category = str(row_cells[0].value).strip()

            # if current_category != "Engineering Inputs":
            if not current_category:
                continue

            param_name = str(row_cells[1].value).strip() if row_cells[1].value else ""

            if not param_name:
                continue

            if verbose:
                print(f" → Processing parameter: {param_name}")

            key = normalize(param_name)

            if key not in param_value_map:
                skipped.append(f"[SKIP] {sheet_name} ({current_category}): parameter '{param_name}' not found in datasheets")
                continue

            value = param_value_map[key]

            # Write value to all unit columns (D onward)
            for col in range(4, max_col + 1):
                if ws_master.cell(row=row_cells[0].row, column=col).value in (None,""):
                    ws_master.cell(row=row_cells[0].row, column=col).value = value

            if verbose:
                print(f" → Wrote '{value}' to all units for '{param_name}'")

    output = BytesIO()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Master_DataSheet_EngineeringInputsPopulated_{timestamp}.xlsx"
    wb_master.save(output)
    output.seek(0)

    return output, filename, skipped


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 3:
        print("Usage: python populate_engineering_inputs.py <master.xlsx> <datasheets.xlsx>")
        sys.exit(1)

    master_path = sys.argv[1]
    datasheets_path = sys.argv[2]

    with open(master_path, "rb") as mf, open(datasheets_path, "rb") as df:
        output, filename, skipped = populate_engineering_inputs(mf, df, verbose=True)

        with open(filename, "wb") as out_f:
            out_f.write(output.read())

        print(f"\n✅ Engineering Inputs populated. Output file: {filename}")
        if skipped:
            print("\n⚠️ Skipped parameters:")
            for s in skipped:
                print("  -", s)
