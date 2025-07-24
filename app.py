"""
app.py

Streamlit frontend for automating the generation and population of a Master Equipment Datasheet.

Steps:
1️⃣ Upload a raw .xlsm file with multiple equipment sheets.
2️⃣ Generate a categorized master datasheet with grouped input sections.
3️⃣ Upload a SysCAD streamtable Excel file to populate SysCAD Inputs into the master datasheet.
4️⃣ Upload datasheets workbook to populate Engineering Inputs.
5️⃣ Download final Excel file with all populated data.

Author: Asfiya Khanam
Updated: July 2025
"""

from io import BytesIO
import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
import zipfile

from automation_test1 import generate_master_datasheet
from populate_equipment_names import populate_equipment_names
from populate_parameters import populate_parameters
from populate_engineering_inputs import populate_engineering_inputs

# ------------------------
# Function to split final master into separate Excel files
# ------------------------
def split_workbook_by_sheet(master_bytes_io):
    """
    Splits a populated master workbook into individual Excel files per sheet.
    Returns a BytesIO stream containing a ZIP archive.
    """
    zip_buffer = BytesIO()
    wb_full = load_workbook(master_bytes_io, data_only=True)

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        for sheet_name in wb_full.sheetnames:
            ws_source = wb_full[sheet_name]

            wb_single = Workbook()
            ws_new = wb_single.active
            ws_new.title = sheet_name

            for row in ws_source.iter_rows(values_only=True):
                ws_new.append(row)

            stream = BytesIO()
            wb_single.save(stream)
            stream.seek(0)

            safe_name = sheet_name.replace("/", "_").replace("\\", "_")
            zipf.writestr(f"{safe_name}.xlsx", stream.read())

    zip_buffer.seek(0)
    return zip_buffer

# ------------------------
# Streamlit App
# ------------------------

st.title("📄 Master Equipment Datasheet Automation Tool")

verbose = False  # Hidden toggle for debugging if needed

st.markdown("""
This tool helps you:
1. Generate a clean, categorized master datasheet from the Excel datasheets workbook.
2. Populate Equipment names.
3. Populate SysCAD parameter values.
4. Populate Engineering Inputs.
""")

# ------------------------
# Step 1: Generate Master Sheet
# ------------------------
st.header("Step 1: Generate Master Datasheet")
st.markdown("""
**What happens in this step?**
- Extracts equipment-wise parameters from your datasheet file (Col I).
- Groups them under 5 categories:
    - SysCAD Inputs
    - Engineering Inputs
    - Lab/Pilot Inputs
    - Project Constants
    - Vendor Inputs
- Creates one formatted sheet per equipment.
""")

uploaded_raw = st.file_uploader("Upload your raw equipment .xlsm file", type=["xlsm"])
if uploaded_raw and st.button("Generate Master Sheet"):
    raw_bytes = uploaded_raw.read()
    output_stream, output_filename = generate_master_datasheet(BytesIO(raw_bytes))
    output_stream.seek(0)

    st.session_state["generated_master"] = output_stream
    st.session_state["raw_datasheets_workbook"] = raw_bytes

    st.success("✅ Master datasheet has been successfully generated!")
    st.download_button(
        label="📥 Download Master Sheet",
        data=output_stream,
        file_name=output_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ------------------------
# Step 2: Populate Equipment Names
# ------------------------
st.header("Step 2: Populate Equipment Names")
st.markdown("""
**What happens in this step?**
- Reads equipment names from the SysCAD report - Detailed Streamtable, Equipment & Stream List sheet.
- Maps equipment codes explicitly.
- Automatically generates implied equipment where applicable.
""")

use_generated_step2 = False
if "generated_master" in st.session_state:
    use_generated_step2 = st.radio(
        "Choose master sheet to populate:",
        ["Use the one generated in Step 1", "Upload a different master sheet"],
        key="step2_radio"
    ) == "Use the one generated in Step 1"

if use_generated_step2:
    master_bytes_step2 = st.session_state["generated_master"]
else:
    uploaded_master_step2 = st.file_uploader("Upload the master sheet", type=["xlsx"], key="master2")
    if uploaded_master_step2:
        master_bytes_step2 = BytesIO(uploaded_master_step2.read())
    else:
        master_bytes_step2 = None

uploaded_stream_step2 = st.file_uploader("Upload the detailed streamtable", type=["xlsx"], key="stream2")

if master_bytes_step2 and uploaded_stream_step2 and st.button("Populate Equipment Names"):
    stream_bytes_step2 = BytesIO(uploaded_stream_step2.read())
    result_step2, filename_step2, skipped_step2 = populate_equipment_names(
        master_bytes_step2, stream_bytes_step2, verbose=verbose
    )

    if skipped_step2:
        st.warning("⚠️ Some equipment were not matched or had issues:")
        st.text_area("Skipped Items", "\n".join(skipped_step2), height=200)

        skipped_df = pd.DataFrame(skipped_step2, columns=["Skipped"])
        skipped_csv = skipped_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="📥 Download Skipped Items List",
            data=skipped_csv,
            file_name="skipped_equipment.csv",
            mime="text/csv"
        )

    st.success("✅ Equipment names populated successfully.")
    st.session_state["master_with_equipment"] = result_step2
    st.session_state["uploaded_stream_content"] = uploaded_stream_step2.getvalue()

    st.download_button(
        label="📥 Download Populated Master Sheet",
        data=result_step2,
        file_name=filename_step2,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ------------------------
# Step 3: Populate Parameters
# ------------------------
st.header("Step 3: Populate Parameters")
st.markdown("""
**What happens in this step?**
- Uses the populated master sheet (with equipment names from Step 2).
- Reads equipment & stream tags from your detailed streamtable.
- Looks up stream data in **Stream Table V**.
- Maps & writes parameters into the master datasheet under **SysCAD Inputs** section.
""")

use_generated_step3 = False
if "master_with_equipment" in st.session_state:
    use_generated_step3 = st.radio(
        "Choose master sheet to populate:",
        ["Use the one generated in Step 2", "Upload a different master sheet"],
        key="step3_radio"
    ) == "Use the one generated in Step 2"

if use_generated_step3:
    master_bytes = st.session_state["master_with_equipment"]
else:
    uploaded_master = st.file_uploader("Upload the master sheet", type=["xlsx"], key="master3")
    if uploaded_master:
        master_bytes = BytesIO(uploaded_master.read())
    else:
        master_bytes = None

if "uploaded_stream_content" in st.session_state:
    stream_bytes = BytesIO(st.session_state["uploaded_stream_content"])
else:
    stream_bytes = None

if master_bytes and stream_bytes and st.button("Populate Parameters"):
    result, filename, skipped = populate_parameters(master_bytes, stream_bytes, verbose=verbose)

    if skipped:
        st.warning("⚠️ Some equipment or streams could not be populated.")
        st.text_area("Skipped Items", "\n".join(skipped), height=200)

        skipped_df = pd.DataFrame(skipped, columns=["Skipped"])
        skipped_csv = skipped_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="📥 Download Skipped Items List",
            data=skipped_csv,
            file_name="skipped_parameters.csv",
            mime="text/csv"
        )

    st.success("✅ Parameters populated successfully into master sheet.")
    st.session_state["master_with_parameters"] = result

    st.download_button(
        label="📥 Download Populated Master Sheet",
        data=result,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ---------------------------------------------------------
# Step 4: Populate Engineering Inputs and Project Constants
# ---------------------------------------------------------
st.header("Step 4: Populate Engineering Inputs and Project Constants")
st.markdown("""
**What happens in this step?**
- Reads the parameters under **Engineering Inputs**  and **Project Constants** in the master sheet.
- Fetches values from the datasheets workbook (col K).
- Populates them for each unit in the master sheet.
""")

use_generated_step4 = False
if "master_with_parameters" in st.session_state:
    use_generated_step4 = st.radio(
        "Choose master sheet to populate:",
        ["Use the one generated in Step 3", "Upload a different master sheet"],
        key="step4_radio"
    ) == "Use the one generated in Step 3"

if use_generated_step4:
    master_bytes_step4 = st.session_state["master_with_parameters"]
else:
    uploaded_master_step4 = st.file_uploader("Upload the master sheet", type=["xlsx"], key="master4")
    if uploaded_master_step4:
        master_bytes_step4 = BytesIO(uploaded_master_step4.read())
    else:
        master_bytes_step4 = None

use_existing_datasheet = False
if "raw_datasheets_workbook" in st.session_state:
    use_existing_datasheet = st.radio(
        "Choose datasheets workbook:",
        ["Use the one uploaded in Step 1", "Upload a different datasheet workbook"],
        key="step4_datasheet_radio"
    ) == "Use the one uploaded in Step 1"

if use_existing_datasheet:
    datasheet_bytes = BytesIO(st.session_state["raw_datasheets_workbook"])
else:
    uploaded_datasheet = st.file_uploader("Upload the datasheets workbook", type=["xls", "xlsx", "xlsm"], key="datasheets")
    datasheet_bytes = BytesIO(uploaded_datasheet.read()) if uploaded_datasheet else None

if master_bytes_step4 and datasheet_bytes and st.button("Populate Engineering Inputs"):
    result_step4, filename_step4, skipped_step4 = populate_engineering_inputs(
        master_bytes_step4, datasheet_bytes, verbose=verbose
    )

    if skipped_step4:
        st.warning("⚠️ Some parameters could not be matched:")
        st.text_area("Skipped Parameters", "\n".join(skipped_step4), height=200)

    st.success("✅ Engineering Inputs populated successfully.")
    st.session_state["master_with_engineering_inputs"] = result_step4


    st.download_button(
        label="📥 Download Master Sheet with Engineering Inputs",
        data=result_step4,
        file_name=filename_step4,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ------------------------
# Step 4.5: Split Master into Separate Equipment Files
# ------------------------
st.header("Step 4.5: Export Individual Equipment Files")
st.markdown("""
**What happens in this step?**
- Takes the populated master sheet (from Step 4).
- Creates one Excel file per equipment sheet.
- Bundles all files into a ZIP archive for download.
""")

if "master_with_engineering_inputs" in st.session_state:
    zip_file = split_workbook_by_sheet(BytesIO(
        st.session_state["master_with_engineering_inputs"].getvalue()
    ))

    st.download_button(
        label="📥 Download Equipment Sheets (ZIP)",
        data=zip_file,
        file_name="PopulatedSheets_SplitByEquipment.zip",
        mime="application/zip"
    )
else:
    st.info("Complete Step 4 first to enable export of split files.")

    # if "master_with_engineering_inputs" in st.session_state:
    #     zip_file = split_workbook_by_sheet(BytesIO(st.session_state["master_with_engineering_inputs"].getvalue()))
    #     st.download_button(
    #         label="📥 Download Equipment Sheets (ZIP)",
    #         data=zip_file,
    #         file_name="PopulatedSheets_SplitByEquipment.zip",
    #         mime="application/zip"
    # )

